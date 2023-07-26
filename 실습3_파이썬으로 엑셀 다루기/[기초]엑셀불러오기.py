# -*- coding: euc-kr -*-
import openpyxl

#기존 엑셀파일이 저장된 경로
fpath = r"C:\pratice_crolling\실습3_파이썬으로 엑셀 다루기\파일이름.xlsx"

# 1) 엑셀 파일 불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트선택
ws = wb['워크시트이름']

# 3) 데이터 수정하기 (엑셀만들기.py와 똑같음)
ws['A3'] = 456
ws['B3'] = '성기훈'

# 4) 수정 후 엑셀 저장하기
wb.save(fpath)
