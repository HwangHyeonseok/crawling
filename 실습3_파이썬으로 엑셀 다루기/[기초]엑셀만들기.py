# -*- coding: euc-kr -*-
import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('워크시트이름')

# 3) 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 4) 엑셀 저장하기
#wb.save('파일이름.xlsx')

# 4-1) 지정된 곳에 엑셀 저장하기
# r은 \을 사용하기 위해 \\와 같이 하지 않아도 되게끔 한다.
# 즉, 이스케이프 시퀀스를 자동해주는 역할 => r
wb.save(r'C:\pratice_crolling\실습3_파이썬으로 엑셀 다루기\파일이름.xlsx')


