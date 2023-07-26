# -*- coding: euc-kr -*-
#���� reference : https://www.dinolabs.ai/163






#���� : ���� ��ǰ ������ ���� ���� ������ ���ǵ帳�ϴ�.
#���� ���� : �ȳ��ϼ���. K��ũ ȫ�� ������ ���� XXX �Դϴ�.
# ���� ��ǰ �ǸŸ� ���� ���� ��ǰ �����Ͱ� �ʿ��մϴ�.
# �Ʒ� Ű���� ���� ����, �귣���, ��ǰ��, ����, ����������ũ�� ������ ������ �ּ���. (1~100��)
# [���̹� ���콺, ���� Ű����, 27��ġ �����]
# ��, ���� ��ǰ�� �����Ѵ�. �귣����� ������ ����д�.










import requests
from bs4 import BeautifulSoup
import pyautogui
import time
import openpyxl

rank = 1 # ����
cur_page = 1 # ���� ������

# ���� ���� ����
excel = openpyxl.Workbook()
row = 2 #������ ��


find_product_count = int(pyautogui.prompt("�� ���� ��ǰ�� ã���ðڽ��ϱ�?"))
    # ã������ ��ǰ ������ŭ ũ�Ѹ� ����
for i in range(0,find_product_count,1):
    search = pyautogui.prompt("ã���÷��� ��ǰ�� �̸��� �ۼ����ּ���.")
    ws = excel.create_sheet(f"{search}") # ���� ��Ʈ ����
    #���� ���� �߰�
    #ws.append(['����', '�귣���', '��ǰ��', '����', '��������'])
    ws['A1'] = "����"
    ws['B1'] = "�귣���"
    ws['C1'] = "��ǰ��"
    ws['D1'] = "����"
    ws['E1'] = "��������"

    for cur_page in range(1,5,1):
    # ���� ����Ʈ�� ������Ʈ �Ǹ鼭 User-Agent �θ� ũ�Ѹ��ϴ� ���� �Ұ�������. �׷��� header ����
        main_url = f"https://www.coupang.com/np/search?q={search}&channel=user&component=&eventCategory=SRP&trcid=&traid=&sorter=scoreDesc&minPrice=&maxPrice=&priceRange=&filterType=&listSize=36&filter=&isPriceRange=false&brand=&offerCondition=&rating=0&page={cur_page}&rocketAll=false&searchIndexingToken=1=9&backgroundColor="
        header = {
            'Host': 'www.coupang.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3',
        }
        response = requests.get(main_url, headers=header)
        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        products = soup.select("a.search-product-link") # ��ǰ�� �ش��ϴ� �ڵ���� ����Ʈ�� �����

        for product in products:
            ad = product.select("span.ad-badge-text") # ���� �ִ� �͵��� ����
            bestseller = product.select(".best-seller-search-product-wrap") # ����Ʈ���� ��ǰ�鵵 ����
            if(len(ad) > 0 or len(bestseller) > 0): # ���� �ְų� ����Ʈ���� ��ǰ�� url�� �������� �ʴ´�.
                continue

            product_url = "https://www.coupang.com" + product.attrs['href'] # ���� ���� ��ǰ�� ��ǰ url ��ũ�� �����´�.
            # ��ǰ �� ����Ʈ�� ����
            response = requests.get(product_url, headers=header) 
            html = response.text
            product_page_soup = BeautifulSoup(html, "html.parser")
            # ����, �귣���, ��ǰ��, ����, ����������ũ ����
            try:
                brand = product_page_soup.select_one(".prod-brand-name").text.strip()
            except:
                brand = "�귣�� ����"
            try:
                product_name = product_page_soup.select_one(".prod-buy-header__title").text.strip()
            except:
                product_name = "��ǰ �̸� ����"
            try:
                price = product_page_soup.select_one(".total-price").text.strip()
            except:
                pirce = "���� ���� ����"
            # ����������ũ�� product_url ��ũ

        
            # 100���� ��������� ���� �������´�.
            if(rank > 100):
                break

            # ������ �����͸� ����.
            ws[f'A{row}'] = rank
            ws[f'B{row}'] = brand
            ws[f'C{row}'] = product_name
            ws[f'D{row}'] = price
            ws[f'E{row}'] = product_url
            row+=1

            print(rf"����:{rank} | �귣��:{brand} | ��ǰ��:{product_name} | ����:{price}")
            print(f"��ũ������ : {product_url}")
            rank+=1
            

        
    rank = 1 # �� ���Ͱ� ������ �ٽ� ������ �ʱ�ȭ
    row = 2 # �� ���Ͱ� ������ �ٽ� ���� �ʱ�ȭ

#���� ����
excel.save(r"C:\pratice_crolling\��ȭ3_���� ��ǰ ũ�Ѹ�\S���� ũ�Ѹ� ����.xlsx")