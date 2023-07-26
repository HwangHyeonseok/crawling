# -*- coding: euc-kr -*-

from openpyxl import Workbook

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성
ws = wb.create_sheet("hyeonseok")

# 해당 시트에 셀 데이터 추가하기
ws['A1'] = "현석"

# 엑셀 저장하기
wb.save(r"C:\pratice_crolling\심화1_\test.xlsx")

