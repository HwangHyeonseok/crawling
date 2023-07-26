# -*- coding: euc-kr -*-
# -------------------------------------------------------���� ��Ȳ----------------------------------------------------------
# ��Ʃ�긦 Ű��� ���� ����Դϴ�. ��Ʃ�긦 �����Ű�� ���� �� �ȸ��� ������ �̾Ƴ����� �մϴ�.
# ���� �м��� ���ؼ� "��Ʃ�� ����Ʈ���� �˻��� �Ͽ� ���� 1������ 200�� ������ ����/��ȸ��/��¥"��
# "�ϳ��� ����"�� �������ִ� ũ�Ѹ� �ϴ� ���α׷��� ������ּ���.

# �� �䱸 �����Դϴ�.
# ������ ���� �״�� ����ϸ� �˴ϴ�.
# ��ȸ���� 15��ȸ���, ���ڷ� 150000 �� ������ ����ǵ��� �ϸ� �˴ϴ�.
# ��¥�� 3���� �� �� ���� ���� �������� ������ ���ε��� �Ⱓ�� ������ ����ǵ��� �ϸ� �˴ϴ�.
# ��Ʃ�� ���̺� ����� ��¿��� �������ּ���.

# -------------------------------------------------------���� �м�----------------------------------------------------------
# * 1) ���� ����Ʈ(requests) vs ���� ����Ʈ(selenium) �Ǻ�
# ��Ʃ�� ���������� �˻�� �˻��ϰ� ������ ��ũ���� ������ ������� �߰������� �� �ߴµ�, url �ּҴ� �ٲ��� �ʴ´�.
# ��, ���� ����Ʈ�� �Ǻ��� �� �ְ� selenium ���� ũ�Ѹ� �ϴ� ���� �� ���� ���̶�� �Ǵ��Ͽ���.
# -------------------------------------------------------�ҽ� �ڵ�----------------------------------------------------------

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time

#������� �Է� �ޱ� (�˻��� �Է�)
import pyautogui
search = pyautogui.prompt("��Ʃ�� ũ�Ѹ� ���α׷��Դϴ�. 200���� ������ ����/��ȸ��/��¥�� ũ�Ѹ��մϴ�. �˻�� �Է����ּ���.")

import openpyxl

# 1) ���� �����
wb = openpyxl.Workbook()

# 2) ���� ��ũ��Ʈ �����
ws = wb.create_sheet(search)

# 3) ���� �ʺ� �� ����
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 120
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 10

# 3) ���� ���� �ֱ�
ws['A1'] = "���� ��ȣ"
ws['B1'] = "���� ����"
ws['C1'] = "��ȸ��"
ws['D1'] = "��¥"

# "��ȸ�� 8.3��ȸ -> 830000000 ���� �Ľ��ϱ� ���� �Լ�"
def parse_views(views_str):
    if 'õȸ' in views_str:
        views_str = views_str.replace('��ȸ�� ', '').replace('õȸ', '').strip()
        views = float(views_str) * 1000
    elif '��ȸ' in views_str:
        views_str = views_str.replace('��ȸ�� ', '').replace('��ȸ', '').strip()
        views = float(views_str) * 10000
    elif '��ȸ' in views_str:
        views_str = views_str.replace('��ȸ�� ', '').replace('��ȸ', '').strip()
        views = float(views_str) * 100000000
    elif 'ȸ' in views_str:
        views_str = views_str.replace('��ȸ�� ', '').replace('ȸ', '').strip()
        views = views_str
    else: #��ȸ���� ���� ���
        views = 0

    return int(views)

#������ ���� ���� 
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# ���ʿ��� ���� �޽��� ���ֱ�
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service() # chrome-for-testing ������ ���� ����Ǵ� �ڵ�
browser = webdriver.Chrome(service=service, options=chrome_options)

# �������� �ش� �ּ� �̵�
browser.get(f"https://www.youtube.com/results?search_query={search}")
browser.implicitly_wait(5) # webpage�� �ε� �� ������ �ִ� 5�ʱ��� ��ٷ��ش�. 

# ��ũ�� ������ (��, ���� 200���� �����ϹǷ� 200�������� ������.)
before_height = browser.execute_script("return window.scrollY")

while True:
    browser.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
    time.sleep(2) # 2�� ���
    
    after_height = browser.execute_script("return window.scrollY")
    print("�� :" + str(before_height))
    print("�� :" + str(after_height))

    video_count = len(browser.find_elements(By.CSS_SELECTOR, "div#dismissible.style-scope.ytd-video-renderer"))
    if(video_count >= 200 ): # ���� 200���� ä�� ��� - div �±��� id �Ӽ����� dismissible �̸鼭 ���ÿ� class �Ӽ����� style-scope ytd-video-renderer �� ���� 200���� �Ѵ��� �˻��Ѵ�.
        print("���� 200���� ã�ҽ��ϴ�.")
        break
    
    elif before_height == after_height: # ��ũ���� ������ �� ���� ���
        print(f"�˻��� �� �ִ� ������ {video_count}�� �Դϴ�. ��ũ���� �� ���Ƚ��ϴ�.")
        break



    before_height = after_height

infos = browser.find_elements(By.CSS_SELECTOR, "div.text-wrapper.style-scope.ytd-video-renderer")

#nth-child : ��� �ڽ��� �������� ã��
#nth-of-type: �ش��ϴ� �ڽ� �±� ��ҿ����� ������ ã��

video_index = 1 #���� ���� ���� ���� ���� (n��° ����)
# info �ڵ� �ȿ��� ��ȸ�ϸ鼭 ����/��ȸ��/��¥ ���� ������ ���´�.
for info in infos:
    # ����
    subject = info.find_element(By.CSS_SELECTOR, "a#video-title") # ������� ����Ʈ�� ��´�.
    view = info.find_element(By.CSS_SELECTOR, "div#metadata-line>span:nth-of-type(1)") # ��ȸ�� - div �±��� id �Ӽ����� metadata-line�� �ٷ� �ڽ��� span �±��� ù ��° ���� �����´�.
    view_num = parse_views(view.text) # ��ȸ�� �Ľ� ex. ��ȸ�� 16��ȸ -> 160,000
    try:
        date = info.find_element(By.CSS_SELECTOR, "div#metadata-line>span:nth-of-type(2)") # ��¥
    except:
        print("����� ���̺��Դϴ�.")
        continue
    
    print(f"{video_index}��° ���� - ���� : {subject.text} ��ȸ�� : {view_num} ��¥ : {date.text}")
    # ������ ������ ����
    ws[f'A{video_index+1}'] = video_index
    ws[f'B{video_index+1}'] = subject.text
    ws[f'C{video_index+1}'] = view_num
    ws[f'D{video_index+1}'] = date.text
    
    video_index += 1

# ���� �����ϱ�
wb.save(rf'C:\pratice_crolling\��ȭ6_��Ʃ�� ũ�Ѹ�\{search}.xlsx')
print("ũ�Ѹ� ���� ���� ���� �� ���� �Ϸ�")